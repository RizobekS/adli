from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Iterable

from django.conf import settings
from django.db.models import DateTimeField, OuterRef, Q, Subquery
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from apps.agency.models import Department, Employee as AgencyEmployee
from apps.requests.models import Request, RequestHistory


GROUP_HEAD_OF_DEPARTMENT = "head_of_department"
REPORT_TEMPLATE_PATH = Path(settings.BASE_DIR) / "static" / "report_templates" / "overdue_requests.xlsx"
REPORT_COLUMN_COUNT = 10
REPORT_DATE_FORMAT = "dd.mm.yyyy"


@dataclass(frozen=True)
class OverdueReportRow:
    request_id: int
    request_public_id: str
    department_name: str
    number_received: str
    work_started_date: date | None
    sender: str
    content: str
    responsible: str
    due_date: date | None
    status_label: str
    overdue_days: int | None
    reason: str
    detail_url_pk: int
    is_unassigned: bool
    is_due_overdue: bool


def _local_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.date()
    return None


def _format_uz_report_date(value: date) -> str:
    months = {
        1: "Yanvar",
        2: "Fevral",
        3: "Mart",
        4: "Aprel",
        5: "May",
        6: "Iyun",
        7: "Iyul",
        8: "Avgust",
        9: "Sentabr",
        10: "Oktabr",
        11: "Noyabr",
        12: "Dekabr",
    }
    return f"{value.year}-yil {value.day}-{months[value.month]}"


def _first_history_subquery(action: str) -> Subquery:
    return Subquery(
        RequestHistory.objects.filter(request_id=OuterRef("pk"), action=action)
        .order_by("created_at")
        .values("created_at")[:1],
        output_field=DateTimeField(),
    )


def _department_head_map() -> dict[int, AgencyEmployee]:
    heads = (
        AgencyEmployee.objects.select_related("user", "department", "position")
        .filter(
            is_active=True,
            department__isnull=False,
            user__groups__name=GROUP_HEAD_OF_DEPARTMENT,
        )
        .order_by("department_id", "position__name", "last_name", "first_name", "user__username")
    )

    result: dict[int, AgencyEmployee] = {}
    for employee in heads:
        if employee.department_id not in result:
            result[employee.department_id] = employee
    return result


def _sender_text(request_obj: Request) -> str:
    company = request_obj.company
    lines = [company.name]
    if company.inn:
        lines.append(f"INN: {company.inn}")
    if company.region_id or company.district_id:
        region = str(company.region) if company.region_id else ""
        district = str(company.district) if company.district_id else ""
        place = ", ".join(part for part in [region, district] if part)
        if place:
            lines.append(place)

    employee = request_obj.employee
    if employee:
        full_name = " ".join(
            part.strip()
            for part in [employee.last_name or "", employee.first_name or "", employee.middle_name or ""]
            if part and part.strip()
        )
        if full_name:
            lines.append(f"Murojaatchi: {full_name}")
        if employee.phone:
            lines.append(f"Tel: {employee.phone}")
        if employee.email:
            lines.append(f"Email: {employee.email}")

    phones = list(getattr(company, "phones", []).all()) if hasattr(company, "phones") else []
    primary_phone = next((phone.phone for phone in phones if phone.is_primary), phones[0].phone if phones else "")
    if primary_phone and not any(primary_phone in line for line in lines):
        lines.append(f"Kompaniya tel.: {primary_phone}")

    return "\n".join(lines)


def _status_label(request_obj: Request, *, report_date: date, is_unassigned: bool) -> str:
    if request_obj.due_date and request_obj.due_date < report_date:
        return "Bajarilmagan"
    if is_unassigned:
        return "Ijrochi belgilanmagan"
    if request_obj.status == Request.Status.WAITING:
        return "Javob kutilmoqda"
    return "Bajarilmoqda"


def overdue_requests_queryset(*, report_date: date, department: Department | None = None):
    first_assigned_at = _first_history_subquery(RequestHistory.Action.ASSIGNED)
    first_step_at = _first_history_subquery(RequestHistory.Action.STEP_ADDED)

    queryset = (
        Request.objects.select_related(
            "company",
            "company__region",
            "company__district",
            "employee",
            "assigned_department",
            "assigned_employee",
            "assigned_employee__department",
            "assigned_employee__position",
            "problem_direction",
        )
        .prefetch_related("company__phones", "directions")
        .exclude(status__in=[Request.Status.DONE, Request.Status.CANCELLED])
        .filter(assigned_department__isnull=False)
        .annotate(first_assigned_at=first_assigned_at, first_step_at=first_step_at)
    )

    if department is not None:
        queryset = queryset.filter(assigned_department=department)

    unassigned_cutoff = report_date - timedelta(days=5)
    return queryset.filter(
        Q(due_date__lt=report_date)
        | Q(
            assigned_employee__isnull=True,
            first_assigned_at__date__lt=unassigned_cutoff,
        )
    ).order_by("assigned_department__name", "due_date", "created_at", "public_id")


def build_overdue_report_rows(*, report_date: date | None = None, department: Department | None = None) -> list[OverdueReportRow]:
    report_date = report_date or timezone.localdate()
    head_map = _department_head_map()
    rows: list[OverdueReportRow] = []

    for request_obj in overdue_requests_queryset(report_date=report_date, department=department):
        first_assigned_date = _local_date(getattr(request_obj, "first_assigned_at", None))
        first_step_date = _local_date(getattr(request_obj, "first_step_at", None))

        is_due_overdue = bool(request_obj.due_date and request_obj.due_date < report_date)
        unassigned_days = (
            (report_date - first_assigned_date).days
            if not request_obj.assigned_employee_id and first_assigned_date
            else None
        )
        is_unassigned = bool(unassigned_days is not None and unassigned_days > 5)

        if not is_due_overdue and not is_unassigned:
            continue

        if request_obj.assigned_employee_id:
            responsible = request_obj.assigned_employee.display_name
            work_started_date = first_step_date or first_assigned_date
        else:
            head = head_map.get(request_obj.assigned_department_id)
            responsible = head.display_name if head else "Rahbar topilmadi"
            work_started_date = first_step_date

        overdue_days = None
        reasons: list[str] = []
        if is_due_overdue and request_obj.due_date:
            overdue_days = (report_date - request_obj.due_date).days
            reasons.append("Ijro muddati o'tgan")
        if is_unassigned and unassigned_days is not None:
            overdue_days = max(overdue_days or 0, unassigned_days)
            reasons.append("Ijrochi 5 kundan ortiq belgilanmagan")

        created_date = _local_date(request_obj.created_at)
        number_received = request_obj.public_id or str(request_obj.pk)
        if created_date:
            number_received = f"{number_received}\n{created_date.isoformat()}"

        content = (request_obj.description or "").strip()
        if request_obj.problem_direction_id:
            content = f"{request_obj.problem_direction}\n{content}" if content else str(request_obj.problem_direction)

        rows.append(
            OverdueReportRow(
                request_id=request_obj.pk,
                request_public_id=request_obj.public_id or str(request_obj.pk),
                department_name=str(request_obj.assigned_department),
                number_received=number_received,
                work_started_date=work_started_date,
                sender=_sender_text(request_obj),
                content=content,
                responsible=responsible,
                due_date=request_obj.due_date,
                status_label=_status_label(request_obj, report_date=report_date, is_unassigned=is_unassigned),
                overdue_days=overdue_days,
                reason="; ".join(reasons),
                detail_url_pk=request_obj.pk,
                is_unassigned=is_unassigned,
                is_due_overdue=is_due_overdue,
            )
        )

    return rows


def summarize_overdue_rows(rows: Iterable[OverdueReportRow]) -> dict[str, int]:
    row_list = list(rows)
    return {
        "total": len(row_list),
        "due_overdue": sum(1 for row in row_list if row.is_due_overdue),
        "unassigned": sum(1 for row in row_list if row.is_unassigned),
        "departments": len({row.department_name for row in row_list}),
    }


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col_idx in range(1, REPORT_COLUMN_COUNT + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def _set_data_style(ws, row_idx: int, *, red_status: bool) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_font = Font(name="Times New Roman", size=12, bold=True, color="000000")
    red_font = Font(name="Times New Roman", size=12, bold=True, color="FF0000")

    for col_idx in range(1, REPORT_COLUMN_COUNT + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.border = border
        cell.font = red_font if red_status and col_idx in (7, 8, 9, 10) else base_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row_idx, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row_idx, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 78


def _prepare_sheet(ws, *, report_date: date) -> None:
    title = (
        f"{_format_uz_report_date(report_date)} holatiga umumiy ijrodagi muddati o'tib "
        "elektron dasturdan yechilmagan murojaatlar bo'yicha ma'lumot"
    )
    headers = [
        "t/r",
        "Ariza raqami va kelib tushgan sanasi",
        "Ariza bo'yicha ish boshlangan sanasi",
        "Jo'natuvchi",
        "Ariza mazmuni",
        "Mas'ul ijrochi",
        "Ijro muddati",
        "Ijro holati",
        "Kechikkan kunlar",
        "Sabab",
    ]

    if ws.max_row > 4:
        ws.delete_rows(5, ws.max_row - 4)

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row in (2, 4):
            ws.unmerge_cells(str(merged_range))

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=REPORT_COLUMN_COUNT)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=REPORT_COLUMN_COUNT)

    ws.cell(2, 1).value = title
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(3, col_idx).value = header

    widths = [7, 23, 21, 36, 58, 22, 16, 19, 15, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 54
    ws.row_dimensions[4].height = 31
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_overdue_report_workbook(*, rows: list[OverdueReportRow], report_date: date) -> bytes:
    workbook = load_workbook(REPORT_TEMPLATE_PATH)
    ws = workbook.active
    ws.title = "Overdue requests"
    _prepare_sheet(ws, report_date=report_date)

    if not rows:
        ws.cell(4, 1).value = "Ma'lumot topilmadi"
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    current_row = 4
    serial = 1
    grouped: dict[str, list[OverdueReportRow]] = {}
    for row in rows:
        grouped.setdefault(row.department_name, []).append(row)

    for dept_idx, (department_name, department_rows) in enumerate(grouped.items()):
        if dept_idx > 0:
            ws.insert_rows(current_row)
            _copy_row_style(ws, 4, current_row)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=REPORT_COLUMN_COUNT)

        ws.cell(current_row, 1).value = department_name
        current_row += 1

        for report_row in department_rows:
            ws.insert_rows(current_row)
            values = [
                serial,
                report_row.number_received,
                report_row.work_started_date,
                report_row.sender,
                report_row.content,
                report_row.responsible,
                report_row.due_date,
                report_row.status_label,
                report_row.overdue_days,
                report_row.reason,
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(current_row, col_idx).value = value

            if report_row.work_started_date:
                ws.cell(current_row, 3).number_format = REPORT_DATE_FORMAT
            if report_row.due_date:
                ws.cell(current_row, 7).number_format = REPORT_DATE_FORMAT

            _set_data_style(ws, current_row, red_status=report_row.is_due_overdue)
            serial += 1
            current_row += 1

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
