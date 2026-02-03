import re
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.companies.models import (
    Company, Category, Region, Direction, Unit,
    CompanyPhone, Position, EmployeeCompany, CompanyDirectionStat
)


SILK_CATEGORY_NAME = "Шелковая промышленность"
DEFAULT_DIRECTOR_POSITION = "Директор"
TOTAL_DIRECTION_TITLE = "JAMI sanoat mahsulotlari"


def norm_inn(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    s = s.replace(".0", "").strip()
    s = re.sub(r"\D+", "", s)
    return s or None


def clean_company_name(s: str) -> str:
    s = (s or "").strip()
    # убираем кавычки типа "Bustan Silk" MCHJ
    s = s.strip('“”"\'')
    return s


def norm_phone(p) -> str | None:
    if not p:
        return None
    digits = re.sub(r"\D+", "", str(p))
    if len(digits) == 9:
        return "+998" + digits
    if len(digits) == 12 and digits.startswith("998"):
        return "+" + digits
    if len(digits) >= 7:
        return digits
    return None


def to_decimal(v) -> Decimal | None:
    """
    Excel обычно отдаёт float. Иногда там 'x' или None.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().lower()
        if not s or s == "x":
            return None
        # на случай если кто-то принёс запятую
        s = s.replace(",", ".")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None

    # float/int
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def get_region_by_code(code: str) -> Region | None:
    return Region.objects.filter(code=str(code).strip()).first()


def get_unit(unit_name: str) -> Unit | None:
    u = (unit_name or "").strip()
    if not u:
        return None
    # сначала по name, потом по short_name
    return (
        Unit.objects.filter(name__iexact=u).first()
        or Unit.objects.filter(short_name__iexact=u).first()
    )


def get_direction(category: Category, title: str) -> Direction:
    t = (title or "").strip()
    obj = Direction.objects.filter(category=category, title__iexact=t).first()
    if obj:
        return obj
    return Direction.objects.create(category=category, title=t)


def is_region_row(cell_c) -> bool:
    if not cell_c or not isinstance(cell_c, str):
        return False
    # формат: "1735 - Qoraqalpog‘iston Respublikasi"
    return bool(re.match(r"^\s*\d+\s*-\s*.+", cell_c))


def parse_region_code(cell_c: str) -> str:
    m = re.match(r"^\s*(\d+)\s*-\s*(.+)", cell_c.strip())
    return m.group(1) if m else ""


def is_company_row(cell_b, cell_e) -> bool:
    # В колонке B часто "1.1", "2.3" (код строки)
    # В колонке E ИНН (число)
    if cell_e is None:
        return False
    inn = norm_inn(cell_e)
    if not inn:
        return False
    # Хедер тоже имеет текст в E, но там "Korxona INN raqami"
    # Отсечём нечисловые и слишком короткие.
    if len(inn) < 7:
        return False
    return True


class Command(BaseCommand):
    help = "Import silk companies from XLSX with per-direction stats for 2025/2026."

    def add_arguments(self, parser):
        parser.add_argument("--path", type=str, required=True)
        parser.add_argument("--sheet", type=str, default="Sheet1")
        parser.add_argument("--dry-run", action="store_true", default=False)

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["path"]
        sheet_name = opts["sheet"]
        dry_run = opts["dry_run"]

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        silk_cat, _ = Category.objects.get_or_create(name=SILK_CATEGORY_NAME)
        director_pos, _ = Position.objects.get_or_create(name=DEFAULT_DIRECTOR_POSITION)
        total_dir = get_direction(silk_cat, TOTAL_DIRECTION_TITLE)

        current_region: Region | None = None
        current_company: Company | None = None

        created_companies = 0
        updated_companies = 0
        created_stats = 0
        updated_stats = 0

        # важные колонки по твоему файлу:
        # C=3: название/направление/регион
        # D=4: unit
        # E=5: inn
        # F=6: employees
        # G=7: director
        # H=8: phone
        # I=9: 2025 qty
        # J=10: 2025 volume
        # K=11: 2026 qty
        # L=12: 2026 volume

        for r in range(1, ws.max_row + 1):
            c = ws.cell(r, 3).value  # C
            d = ws.cell(r, 4).value  # D
            e = ws.cell(r, 5).value  # E
            f = ws.cell(r, 6).value  # F
            g = ws.cell(r, 7).value  # G
            h = ws.cell(r, 8).value  # H
            i = ws.cell(r, 9).value  # I
            j = ws.cell(r, 10).value # J
            k = ws.cell(r, 11).value # K
            l = ws.cell(r, 12).value # L
            b = ws.cell(r, 2).value  # B (код)

            # 1) Регион
            if is_region_row(c):
                code = parse_region_code(c)
                current_region = get_region_by_code(code)
                current_company = None
                continue

            # 2) Компания
            if is_company_row(b, e):
                inn = norm_inn(e)
                name = clean_company_name(str(c or "").strip())
                if not inn or not name:
                    continue

                # employees
                jobs = None
                if f is not None:
                    try:
                        jobs = int(float(f))
                    except Exception:
                        jobs = None

                company, created = Company.objects.update_or_create(
                    inn=inn,
                    defaults={
                        "name": name,
                        "region": current_region,
                        "number_of_jobs": jobs,
                    }
                )
                current_company = company

                if created:
                    created_companies += 1
                else:
                    updated_companies += 1

                # категория: оставляем старую основную, но добавляем шелк в M2M
                company.categories.add(silk_cat)
                if company.category_id is None:
                    company.category = silk_cat
                    company.save(update_fields=["category"])

                # директор (как сотрудник)
                director_name = (str(g or "").strip() or "")
                if director_name:
                    # примитивное разбиение ФИО: LAST FIRST MIDDLE...
                    parts = director_name.split()
                    last_name = parts[0] if len(parts) >= 1 else None
                    first_name = parts[1] if len(parts) >= 2 else None
                    middle_name = " ".join(parts[2:]) if len(parts) >= 3 else None

                    EmployeeCompany.objects.get_or_create(
                        company=company,
                        position=director_pos,
                        last_name=last_name,
                        first_name=first_name,
                        middle_name=middle_name,
                    )

                # телефон компании
                phone = norm_phone(h)
                if phone:
                    CompanyPhone.objects.get_or_create(company=company, phone=phone)

                # общий объём компании (строка компании): сохраняем как TOTAL_DIRECTION_TITLE
                v2025 = to_decimal(j)
                v2026 = to_decimal(l)

                # qty там часто "x", так что quantity=None
                for year, vol in ((2025, v2025), (2026, v2026)):
                    if vol is None:
                        continue
                    stat, s_created = CompanyDirectionStat.objects.update_or_create(
                        company=company,
                        direction=total_dir,
                        year=year,
                        defaults={
                            "unit": None,
                            "quantity": None,
                            "volume_bln_sum": vol,
                        }
                    )
                    created_stats += int(s_created)
                    updated_stats += int(not s_created)

                continue

            # 3) Направления компании (подстроки после компании)
            if current_company and isinstance(c, str) and c.strip():
                title = c.strip()
                # игнорим строки типа "JAMI sanoat mahsulotlari" и похожие служебные если попадутся
                if title.lower().startswith("jami"):
                    continue

                unit = get_unit(str(d or ""))

                qty2025 = to_decimal(i)
                vol2025 = to_decimal(j)
                qty2026 = to_decimal(k)
                vol2026 = to_decimal(l)

                # если вообще нет данных, пропускаем
                if qty2025 is None and vol2025 is None and qty2026 is None and vol2026 is None:
                    continue

                direction = get_direction(silk_cat, title)

                # добавим направление в M2M компании для фильтров
                current_company.directions.add(direction)

                for year, qty, vol in (
                    (2025, qty2025, vol2025),
                    (2026, qty2026, vol2026),
                ):
                    if qty is None and vol is None:
                        continue
                    stat, s_created = CompanyDirectionStat.objects.update_or_create(
                        company=current_company,
                        direction=direction,
                        year=year,
                        defaults={
                            "unit": unit,
                            "quantity": qty,
                            "volume_bln_sum": vol,
                        }
                    )
                    created_stats += int(s_created)
                    updated_stats += int(not s_created)

        if dry_run:
            # откатим транзакцию
            raise SystemExit(
                f"[DRY RUN] companies created={created_companies}, updated={updated_companies}, "
                f"stats created={created_stats}, updated={updated_stats}"
            )

        self.stdout.write(self.style.SUCCESS(
            f"Done. companies created={created_companies}, updated={updated_companies}, "
            f"stats created={created_stats}, updated={updated_stats}"
        ))
